from datetime import datetime, timedelta
from openpyxl import Workbook

from django.shortcuts import render
from django.views.generic.base import View, TemplateResponseMixin
from django.http import HttpResponse

from customers.models import Customer
from organizations.models import CompanyUser


class CustomerListView(TemplateResponseMixin, View):
    template_name = 'customers/list.html'

    def get_date_range(self, range_type):
        today = datetime.now()
        if range_type == "all-time":
            earliest_date = Customer.objects.all().order_by('created_at').first().created_at.date()
            return earliest_date, today
        start_date_methods = {
            "week": lambda: today - timedelta(days=6),
            "month": lambda: today.replace(day=1),
            "quarter": lambda: today.replace(month=((today.month - 1) // 3) * 3 + 1, day=1),
            "half-year": lambda: today.replace(month=1, day=1) if today.month > 6 else today.replace(
                year=today.year - 1, month=7, day=1),
            "year": lambda: today.replace(month=1, day=1)
        }
        start_date = start_date_methods.get(range_type, lambda: None)()
        return start_date, today

    def get(self, request, *args, **kwargs):
        branch_id = request.GET.get('branch', 'all')
        range_type = request.GET.get('range_type', 'month')
        start_date, end_date = self.get_date_range(range_type)

        base_filter = {
            'branch__company__companyuser__user': request.user,
            'created_at__range': (start_date, end_date),
        }
        if branch_id != 'all':
            base_filter['branch_id'] = branch_id
            branch_id = int(branch_id)
        customers = Customer.objects.filter(**base_filter)

        try:
            company_of_current_user = CompanyUser.objects.get(user=request.user).company
            branches = company_of_current_user.branches.all()
        except CompanyUser.DoesNotExist:
            return self.render_to_response({'error': "User is not associated with any company"})

        context = {
            'customers': customers,
            'branch_id': branch_id,
            'range_type': range_type,
            'branches': branches,
        }
        return self.render_to_response(context)


class ExportCustomersView(View):

    def get(self, request, *args, **kwargs):
        # Get filter criteria from request
        branch_id = request.GET.get('branch')
        range_type = request.GET.get('range_type')

        # Apply filters to fetch customers
        customers = self.get_filtered_customers(branch_id, range_type)

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="customers.xlsx"'

        wb = self._create_workbook(customers)
        wb.save(response)
        return response

    def get_filtered_customers(self, branch_id, range_type):
        # Build your filter criteria based on branch_id and range_type
        filters = {}

        if branch_id != 'all':
            filters['branch__id'] = branch_id

        if range_type == 'week':
            # Calculate date range for the past week and add it to filters
            end_date = datetime.now()
            start_date = end_date - timedelta(days=7)
            filters['created_at__range'] = (start_date, end_date)
        elif range_type == 'month':
            # Calculate date range for the past month and add it to filters
            end_date = datetime.now()
            start_date = end_date - timedelta(days=30)
            filters['created_at__range'] = (start_date, end_date)
        elif range_type == 'half-year':
            # Calculate date range for the past six months and add it to filters
            end_date = datetime.now()
            start_date = end_date - timedelta(days=180)
            filters['created_at__range'] = (start_date, end_date)
        elif range_type == 'year':
            # Calculate date range for the past year and add it to filters
            end_date = datetime.now()
            start_date = end_date - timedelta(days=365)
            filters['created_at__range'] = (start_date, end_date)

        # Fetch filtered customers
        customers = Customer.objects.filter(**filters)

        return customers

    def _create_workbook(self, customers):
        wb = Workbook()
        ws = wb.active
        ws.title = "Клиенты"

        # Заголовки столбцов
        columns = ['#', 'Номер телефона', 'Филиал', 'Клиент с']
        for col_num, column_title in enumerate(columns, 1):
            col_letter = chr(64 + col_num)
            ws['{}1'.format(col_letter)] = column_title
            ws.column_dimensions[col_letter].width = 15

        for idx, customer in enumerate(customers, 2):
            ws.cell(row=idx, column=1, value=idx-1)
            ws.cell(row=idx, column=2, value=customer.phone_number)
            ws.cell(row=idx, column=3, value=customer.branch.title)
            ws.cell(row=idx, column=4, value=customer.created_at.strftime('%Y-%m-%d'))

        return wb
